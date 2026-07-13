# -*- coding: utf-8 -*-
"""Generate the data deliverables from kit_schema:
   - 8 styled per-entity XLSX workbooks (Template / Example / Field Guide tabs)
   - 1 master workbook (Read Me / Lookups / All Fields / Legend)
   - CSV templates (machine headers + a separate examples CSV)
   - schema.sql (PostgreSQL DDL)         - schema.json (JSON Schema per entity)
   - validation_rules.json               - import error-report template CSV
"""
import os, json, csv, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import kit_schema as S

OUT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data-import-kit"))
def p(*a): return os.path.join(OUT, *a)
for d in ["", "templates-xlsx", "templates-csv", "schema"]:
    os.makedirs(p(d), exist_ok=True)

# ---- palette ---------------------------------------------------------------
TEAL   = "FF073D52"; TEAL2 = "FF115F7D"; BONE = "FFF3EFE6"; BONE2 = "FFFBF9F4"
WHITE  = "FFFFFFFF"; INK = "FF1B2B31"; AMBER = "FFFFF3D6"; REQ = "FFEAF3F1"
GREY   = "FF6B7A80"; LINE = "FFDDE5E4"; GREENH = "FF0E5A3C"

f_hdr   = Font(name="Calibri", size=11, bold=True, color=WHITE)
f_sub   = Font(name="Calibri", size=10, bold=True, color=WHITE)
f_body  = Font(name="Calibri", size=10, color=INK)
f_help  = Font(name="Calibri", size=9, italic=True, color=GREY)
f_ex    = Font(name="Calibri", size=10, color=INK)
f_title = Font(name="Calibri", size=15, bold=True, color=WHITE)
f_note  = Font(name="Calibri", size=10, color=INK)
f_key   = Font(name="Consolas", size=10, bold=True, color=TEAL2)

fill_hdr = PatternFill("solid", fgColor=TEAL)
fill_sub = PatternFill("solid", fgColor=TEAL2)
fill_req = PatternFill("solid", fgColor=REQ)
fill_ex  = PatternFill("solid", fgColor=AMBER)
fill_bone= PatternFill("solid", fgColor=BONE)
fill_bone2=PatternFill("solid", fgColor=BONE2)
fill_title=PatternFill("solid", fgColor=TEAL)

thin = Side(style="thin", color=LINE)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)

WIDTH = {"id":16,"fk":16,"text":26,"longtext":42,"slug":26,"url":30,"image":32,
         "images":38,"int":13,"decimal":15,"enum":22,"date":14,"geo":13,
         "email":24,"phone":18,"bool":12}
def width(t): return WIDTH.get(t, 20)

def allowed(f):
    if f["enum"]: return S.L(f["enum"])
    if f["type"] in ("int","decimal"): return None
    if f["type"] == "date": return None
    return None

def comment_text(f):
    lines = [f'{f["label_en"]}  |  {f["label_ar"]}',
             f'field: {f["name"]}',
             f'type: {f["type"]}   required: {"YES" if f["required"] else "no"}']
    if f["fk"]:   lines.append(f'links to: {f["fk"]}')
    if f["unique"]: lines.append("must be UNIQUE")
    if f["enum"]: lines.append("allowed: " + ", ".join(S.L(f["enum"])))
    if f["seo"]:  lines.append("SEO field")
    if f["help_en"]: lines.append("• " + f["help_en"])
    if f["help_ar"]: lines.append("• " + f["help_ar"])
    if f["example"]: lines.append("e.g. " + str(f["example"]))
    return "\n".join(lines)

# ---------------------------------------------------------------------------
def add_lists_sheet(wb, fields):
    """Hidden sheet holding vocab columns so dropdowns can reference ranges
    (avoids Excel's 255-char inline-list limit)."""
    ws = wb.create_sheet("_lists")
    ws.sheet_state = "hidden"
    col = 1
    ranges = {}
    seen = {}
    for f in fields:
        if not f["enum"]:
            continue
        if f["enum"] in seen:
            ranges[f["name"]] = seen[f["enum"]]
            continue
        letter = get_column_letter(col)
        vals = S.L(f["enum"])
        ws.cell(row=1, column=col, value=f["enum"]).font = Font(bold=True, size=9, color=GREY)
        for i, v in enumerate(vals, start=2):
            ws.cell(row=i, column=col, value=v)
        rng = f"_lists!${letter}$2:${letter}${1+len(vals)}"
        ranges[f["name"]] = rng
        seen[f["enum"]] = rng
        ws.column_dimensions[letter].width = 24
        col += 1
    return ranges

def style_header_row(ws, fields, row=1):
    ws.row_dimensions[row].height = 30
    for c, f in enumerate(fields, start=1):
        cell = ws.cell(row=row, column=c,
                       value=f["name"] + (" *" if f["required"] else ""))
        cell.font = f_hdr; cell.fill = fill_hdr; cell.alignment = ctr; cell.border = box
        cm = Comment(comment_text(f), "The Village · Data Kit"); cm.width = 320; cm.height = 200
        cell.comment = cm
        ws.column_dimensions[get_column_letter(c)].width = width(f["type"])
    ws.freeze_panes = f"A{row+1}"   # string form: does NOT materialise a phantom cell
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(fields))}{row}"

def apply_validations(ws, fields, ranges, first=3, last=5000):
    for c, f in enumerate(fields, start=1):
        letter = get_column_letter(c)
        dv = None
        if f["enum"]:
            dv = DataValidation(type="list", formula1=ranges[f["name"]], allow_blank=True)
            dv.error = "Pick a value from the drop-down list."
            dv.errorTitle = "Not in list"
            dv.prompt = f'{f["label_en"]} — choose from the list'
        elif f["type"] == "int":
            dv = DataValidation(type="whole", operator="greaterThanOrEqual",
                                formula1="0", allow_blank=True)
            dv.error = "Whole numbers only (0 or more)."
        elif f["type"] == "decimal":
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual",
                                formula1="0", allow_blank=True)
            dv.error = "Numbers only (0 or more). No commas or text."
        elif f["type"] == "date":
            dv = DataValidation(type="date", operator="greaterThanOrEqual",
                                formula1=datetime.date(2000,1,1), allow_blank=True)
            dv.error = "Use a real date, format YYYY-MM-DD."
        if dv is not None:
            dv.add(f"{letter}{first}:{letter}{last}")
            ws.add_data_validation(dv)
        # subtle required tint on the first empty entry rows
        if f["required"]:
            for r in range(first, first+40):
                ws.cell(row=r, column=c).fill = fill_req

def field_guide_sheet(ws, ent):
    heads = ["Field (system name)","Label (EN)","Label (AR)","Required","Type",
             "Allowed values / format","Example","Notes"]
    ws.append(heads)
    ws.row_dimensions[1].height = 26
    for c in range(1, len(heads)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = f_hdr; cell.fill = fill_hdr; cell.alignment = ctr; cell.border = box
    widths = [22,22,22,10,12,34,26,34]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for f in ent["fields"]:
        if f["enum"]:
            av = ", ".join(S.L(f["enum"]))
        elif f["type"] in ("int","decimal"):
            av = "number ≥ 0"
        elif f["type"] == "date":
            av = "date YYYY-MM-DD"
        elif f["type"] == "slug":
            av = "lowercase-with-dashes"
        elif f["type"] in ("image","images"):
            av = "asset path (see naming rules)"
        elif f["type"] == "url":
            av = "https://…"
        else:
            av = "free text"
        notes = []
        if f["pk"]: notes.append("Primary key")
        if f["fk"]: notes.append("Links to " + f["fk"])
        if f["unique"]: notes.append("Unique")
        if f["seo"]: notes.append("SEO")
        row = [f["name"], f["label_en"], f["label_ar"],
               "Required" if f["required"] else "Optional", f["type"], av,
               str(f["example"]), "; ".join(notes)]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(heads)+1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap; cell.border = box; cell.font = f_body
            cell.fill = fill_bone2 if r % 2 == 0 else PatternFill("solid", fgColor=WHITE)
            if c == 1: cell.font = f_key
            if c == 4 and f["required"]:
                cell.font = Font(name="Calibri", size=10, bold=True, color=GREENH)
    ws.freeze_panes = "A2"

def banner(ws, ent, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 4))
    c = ws.cell(row=row, column=1,
        value=f'{ent["num"]}. {ent["title_en"]}  ·  {ent["title_ar"]}   —   '
              f'figures are illustrative; confirm with the developer/advisor')
    c.font = f_title; c.fill = fill_title
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 30

def build_entity_workbook(ent):
    wb = openpyxl.Workbook()
    tmpl = wb.active; tmpl.title = "Template"
    ranges = add_lists_sheet(wb, ent["fields"])
    # Template tab: banner row 1, header row 2, data from row 3
    banner(tmpl, ent, len(ent["fields"]), row=1)
    style_header_row(tmpl, ent["fields"], row=2)
    apply_validations(tmpl, ent["fields"], ranges, first=3, last=5000)
    tmpl.sheet_view.showGridLines = False
    # Example tab
    ex = wb.create_sheet("Example (filled)")
    banner(ex, ent, len(ent["fields"]), row=1)
    style_header_row(ex, ent["fields"], row=2)
    for row in S.SAMPLES[ent["key"]]:
        ex.append(list(row))
        r = ex.max_row
        for c in range(1, len(ent["fields"])+1):
            cell = ex.cell(row=r, column=c)
            cell.fill = fill_ex; cell.alignment = wrap; cell.border = box; cell.font = f_ex
    ex.sheet_view.showGridLines = False
    note = ex.cell(row=ex.max_row+2, column=1,
                   value="↑ Example rows only — do not import this tab. Enter real data on the ‘Template’ tab.")
    note.font = Font(size=10, bold=True, color=TEAL2)
    # Field Guide tab
    fg = wb.create_sheet("Field Guide")
    field_guide_sheet(fg, ent)
    fg.sheet_view.showGridLines = False
    # lock the read-only reference tabs so staff can't break them by accident
    # (the 'Template' tab stays fully editable). No password → trivially removable.
    for ws in (ex, fg):
        ws.protection.sheet = True
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
    lists_ws = wb["_lists"]; lists_ws.protection.sheet = True
    # order tabs
    wb.move_sheet("_lists", offset=len(wb.sheetnames))
    fn = p("templates-xlsx", f'{ent["num"]:02d}_{ent["title_en"].replace(" & ","_").replace(" ","_")}.xlsx')
    wb.save(fn)
    return fn

# ---------------------------------------------------------------------------
def build_master_workbook():
    wb = openpyxl.Workbook()
    # Read Me
    rm = wb.active; rm.title = "Read Me"
    rm.sheet_view.showGridLines = False
    rm.column_dimensions["A"].width = 3
    rm.column_dimensions["B"].width = 110
    def line(txt, font=f_note, h=None):
        rm.append([None, txt]); r = rm.max_row
        rm.cell(row=r, column=2).font = font
        rm.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        if h: rm.row_dimensions[r].height = h
    hd = rm.cell(row=1, column=2, value="THE VILLAGE INVESTMENT — Real-Estate Data Import Kit")
    hd.font = Font(size=16, bold=True, color=TEAL2); rm.row_dimensions[1].height = 26
    line(f'Version {S.BRAND["version"]} · {S.BRAND["date"]} · {S.BRAND["site"]}', f_help)
    line("")
    line("WHAT THIS IS", Font(size=12, bold=True, color=GREENH))
    line("A set of 8 connected spreadsheets your team fills in to describe every area, developer, "
         "project, unit type, unit, amenity, payment plan and joint-venture link. Fill the sheets, "
         "hand them back, and they import cleanly into the database that powers the website.")
    line("")
    line("FILL THE SHEETS IN THIS ORDER (parents before children):", Font(size=12, bold=True, color=GREENH))
    for e in S.ENTITIES:
        line(f'   {e["num"]}.  {e["title_en"]}  ·  {e["title_ar"]}   —   {e["desc_en"]}')
    line("")
    line("GOLDEN RULES", Font(size=12, bold=True, color=GREENH))
    rules = [
        "Each sheet has 3 tabs: ‘Template’ (type here) · ‘Example (filled)’ (a sample) · ‘Field Guide’ (every field explained).",
        "Columns marked with * are REQUIRED. Hover any header (red triangle) to see the full help note.",
        "IDs (Area ID, Developer ID …) may be left blank — they are generated automatically on import. "
        "If you do type them, keep the format shown (e.g. PRJ-00001) and never reuse a code.",
        "To connect rows, copy the exact ID: a Project’s ‘developer_id’ must match a real ‘developer_id’ in the Developers sheet.",
        "Cells with a drop-down arrow only accept a value from the list — pick, don’t type.",
        "Prices are illustrative and must be confirmed with the developer/advisor. We market PRIMARY sale only.",
        "Dates use the format YYYY-MM-DD (e.g. 2030-06-30). Numbers have no commas or currency signs.",
        "Images are entered as a file PATH (e.g. projects/badya/hero.jpg), not pasted pictures. See the Naming guide.",
        "Do not rename tabs, delete the header row, or reorder columns.",
    ]
    for i, r in enumerate(rules, 1):
        line(f'   {i}.  {r}')
    line("")
    line("COLOUR LEGEND", Font(size=12, bold=True, color=GREENH))
    leg = [("Teal header","the column name + * for required", TEAL),
           ("Light-teal cells","a required field — please fill", REQ),
           ("Amber rows","example data (reference only, never imported)", AMBER)]
    for name, meaning, color in leg:
        rm.append([None, f"{name} — {meaning}"]); r = rm.max_row
        rm.cell(row=r, column=2).font = f_note
        rm.cell(row=r, column=1).fill = PatternFill("solid", fgColor=color)
        rm.cell(row=r, column=1).border = box
    line("")
    line("NEED HELP?  Keep a copy of every file before editing, and return the filled files unchanged in structure. "
         "The ‘Data Dictionary.pdf’ and ‘README.pdf’ explain everything in full.", f_help)

    # Lookups
    lk = wb.create_sheet("Lookups")
    lk.sheet_view.showGridLines = False
    lk.cell(row=1, column=1, value="CONTROLLED LISTS — the only accepted values (English is stored; Arabic is a helper)")
    lk.cell(row=1, column=1).font = Font(size=12, bold=True, color=TEAL2)
    col = 1
    for key, pairs in S.LOOKUPS.items():
        h1 = lk.cell(row=3, column=col, value=key); h1.font = f_hdr; h1.fill = fill_hdr; h1.alignment=ctr; h1.border=box
        h2 = lk.cell(row=3, column=col+1, value="عربي"); h2.font = f_hdr; h2.fill = fill_sub; h2.alignment=ctr; h2.border=box
        for i,(en,ar) in enumerate(pairs, start=4):
            a=lk.cell(row=i, column=col, value=en); a.border=box; a.font=f_body; a.alignment=wrap
            b=lk.cell(row=i, column=col+1, value=ar); b.border=box; b.font=f_body; b.alignment=wrap
        lk.column_dimensions[get_column_letter(col)].width = 26
        lk.column_dimensions[get_column_letter(col+1)].width = 22
        col += 3

    # All Fields (full dictionary)
    af = wb.create_sheet("All Fields")
    af.sheet_view.showGridLines = False
    heads = ["Entity","Field","Label (EN)","Label (AR)","Required","Type","Allowed / format","Example","Key"]
    af.append(heads)
    for c in range(1,len(heads)+1):
        cell=af.cell(row=1,column=c); cell.font=f_hdr; cell.fill=fill_hdr; cell.alignment=ctr; cell.border=box
    for i,w in enumerate([20,22,22,22,10,11,34,24,16], start=1):
        af.column_dimensions[get_column_letter(i)].width=w
    for e in S.ENTITIES:
        for f in e["fields"]:
            av = (", ".join(S.L(f["enum"])) if f["enum"] else
                  "number ≥ 0" if f["type"] in ("int","decimal") else
                  "YYYY-MM-DD" if f["type"]=="date" else
                  "lowercase-with-dashes" if f["type"]=="slug" else
                  "asset path" if f["type"] in ("image","images") else
                  "https://…" if f["type"]=="url" else "free text")
            key = "PK" if f["pk"] else ("FK→"+f["fk"].split(".")[0] if f["fk"] else ("unique" if f["unique"] else ""))
            af.append([e["title_en"], f["name"], f["label_en"], f["label_ar"],
                       "Required" if f["required"] else "Optional", f["type"], av, str(f["example"]), key])
            r=af.max_row
            for c in range(1,len(heads)+1):
                cell=af.cell(row=r,column=c); cell.alignment=wrap; cell.border=box; cell.font=f_body
                cell.fill = fill_bone2 if r%2==0 else PatternFill("solid",fgColor=WHITE)
                if c==2: cell.font=f_key
    af.freeze_panes="A2"; af.auto_filter.ref=f"A1:I{af.max_row}"

    fn = p("templates-xlsx", "00_MASTER_Read-me_and_Lookups.xlsx")
    wb.save(fn)
    return fn

# ---------------------------------------------------------------------------
def build_csv():
    for e in S.ENTITIES:
        headers = [f["name"] for f in e["fields"]]
        with open(p("templates-csv", f'{e["num"]:02d}_{e["table"]}.csv'), "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh); w.writerow(headers)
        with open(p("templates-csv", f'{e["num"]:02d}_{e["table"]}__example.csv'), "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh); w.writerow(headers)
            for row in S.SAMPLES[e["key"]]:
                w.writerow(row)
    # error report template
    with open(p("templates-csv", "_import_error_report_TEMPLATE.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["row_number","sheet","column","field","offending_value","severity","error_code","message"])
        w.writerow(["7","projects","developer_id","developer_id","DEV-9999","error","FK_NOT_FOUND",
                    "developer_id 'DEV-9999' does not exist in Developers"])
        w.writerow(["12","units","built_up_area_sqm","built_up_area_sqm","","error","REQUIRED_MISSING",
                    "built_up_area_sqm is required"])
        w.writerow(["20","areas","slug","slug","new cairo","warning","SLUG_FORMAT",
                    "slug should be lowercase-with-dashes; suggested 'new-cairo'"])

# ---------------------------------------------------------------------------
SQLTYPE = {"id":"VARCHAR(20)","fk":"VARCHAR(20)","text":"VARCHAR(255)","longtext":"TEXT",
           "slug":"VARCHAR(160)","url":"VARCHAR(500)","image":"VARCHAR(500)","images":"TEXT",
           "int":"INTEGER","decimal":"NUMERIC(16,2)","enum":"VARCHAR(60)","date":"DATE",
           "geo":"NUMERIC(10,6)","email":"VARCHAR(255)","phone":"VARCHAR(40)","bool":"BOOLEAN"}

def sqltype(f):
    if f["sql"]: return f["sql"]
    if f["name"] in ("latitude","longitude"): return "NUMERIC(10,6)"
    return SQLTYPE.get(f["type"], "VARCHAR(255)")

def build_sql():
    lines = []
    lines.append("-- The Village Investment — Real-Estate schema (PostgreSQL)")
    lines.append(f"-- Generated {S.BRAND['date']} · v{S.BRAND['version']} · scalable to 100,000+ units")
    lines.append("-- Enum-style columns are constrained with CHECK lists so bad values are rejected.\n")
    lines.append("BEGIN;\n")
    for e in S.ENTITIES:
        lines.append(f'-- {e["num"]}. {e["title_en"]} — {e["desc_en"]}')
        lines.append(f'CREATE TABLE IF NOT EXISTS {e["table"]} (')
        col_lines = []
        for f in e["fields"]:
            parts = [f'  {f["name"]:26s} {sqltype(f)}']
            if f["pk"]: parts.append("PRIMARY KEY")
            if f["required"] and not f["pk"]: parts.append("NOT NULL")
            if f["enum"]:
                vals = ", ".join("'" + v.replace("'", "''") + "'" for v in S.L(f["enum"]))
                parts.append(f'CHECK ({f["name"]} IN ({vals}))')
            col_lines.append(" ".join(parts))
        # audit columns
        col_lines.append("  created_at                 TIMESTAMPTZ NOT NULL DEFAULT now()")
        col_lines.append("  updated_at                 TIMESTAMPTZ NOT NULL DEFAULT now()")
        # FKs
        for f in e["fields"]:
            if f["fk"]:
                rt, rc = f["fk"].split(".")
                col_lines.append(f'  CONSTRAINT fk_{e["table"]}_{f["name"]} '
                                 f'FOREIGN KEY ({f["name"]}) REFERENCES {rt}({rc})')
        # unique
        for f in e["fields"]:
            if f["unique"]:
                col_lines.append(f'  CONSTRAINT uq_{e["table"]}_{f["name"]} UNIQUE ({f["name"]})')
        lines.append(",\n".join(col_lines))
        lines.append(");")
        # indexes on FKs + slug
        for f in e["fields"]:
            if f["fk"]:
                lines.append(f'CREATE INDEX IF NOT EXISTS ix_{e["table"]}_{f["name"]} ON {e["table"]}({f["name"]});')
            if f["type"] == "slug":
                lines.append(f'CREATE INDEX IF NOT EXISTS ix_{e["table"]}_{f["name"]} ON {e["table"]}({f["name"]});')
        lines.append("")
    lines.append("COMMIT;")
    open(p("schema", "schema.sql"), "w", encoding="utf-8").write("\n".join(lines))

# ---------------------------------------------------------------------------
JSONTYPE = {"id":"string","fk":"string","text":"string","longtext":"string","slug":"string",
            "url":"string","image":"string","images":"string","int":"integer","decimal":"number",
            "enum":"string","date":"string","geo":"number","email":"string","phone":"string","bool":"boolean"}

def build_json_schema():
    schemas = {}
    for e in S.ENTITIES:
        props, required = {}, []
        for f in e["fields"]:
            sch = {"type": JSONTYPE.get(f["type"], "string"),
                   "title": f["label_en"], "title_ar": f["label_ar"],
                   "description": f["help_en"] or f["label_en"], "examples": [f["example"]]}
            if f["enum"]: sch["enum"] = S.L(f["enum"])
            if f["type"] == "date": sch["format"] = "date"; sch["type"] = "string"
            if f["type"] == "url": sch["format"] = "uri"
            if f["type"] == "email": sch["format"] = "email"
            if f["type"] == "slug": sch["pattern"] = "^[a-z0-9]+(?:-[a-z0-9]+)*$"
            if f["type"] in ("int","decimal","geo"): sch["minimum"] = -1e12 if f["name"] in ("latitude","longitude") else 0
            if f["fk"]: sch["x-foreign-key"] = f["fk"]
            if f["pk"]: sch["x-primary-key"] = True
            if f["unique"]: sch["x-unique"] = True
            props[f["name"]] = sch
            if f["required"]: required.append(f["name"])
        schemas[e["table"]] = {
            "$schema": "https://json-schema.org/draft/2020-12/schema",
            "$id": f'https://{S.BRAND["site"]}/schema/{e["table"]}.json',
            "title": e["title_en"], "title_ar": e["title_ar"], "description": e["desc_en"],
            "type": "object", "additionalProperties": False,
            "properties": props, "required": required,
        }
    bundle = {"$schema":"https://json-schema.org/draft/2020-12/schema",
              "title":"The Village Investment — data import schemas",
              "version":S.BRAND["version"], "generated":S.BRAND["date"],
              "entities":list(schemas.keys()), "definitions":schemas}
    open(p("schema","schema.json"),"w",encoding="utf-8").write(json.dumps(bundle, ensure_ascii=False, indent=2))

# ---------------------------------------------------------------------------
def build_validation_rules():
    rules = {"version":S.BRAND["version"], "generated":S.BRAND["date"],
             "conventions":{"header_row":1,"data_start_row":2,"encoding":"UTF-8",
                            "date_format":"YYYY-MM-DD","decimal_separator":".",
                            "list_separator_multi_value":"|","boolean_values":["Yes","No"]},
             "import_order":[e["table"] for e in S.ENTITIES],
             "id_patterns":{e["table"]: f'^{e["id_prefix"]}-[0-9]{{{4 if e["id_prefix"] in ("AREA","DEV","UTYP","PLAN") else 5}}}[0-9]*$' for e in S.ENTITIES},
             "entities":{}}
    for e in S.ENTITIES:
        frules = []
        for f in e["fields"]:
            r = {"field":f["name"],"required":f["required"],"type":f["type"]}
            if f["enum"]: r["one_of"] = S.L(f["enum"])
            if f["fk"]: r["foreign_key"] = f["fk"]
            if f["unique"]: r["unique"] = True
            if f["pk"]: r["primary_key"] = True; r["auto_generate_if_blank"] = True
            if f["type"] in ("int","decimal","geo"): r["min"] = 0 if f["name"] not in ("latitude","longitude") else -180
            if f["type"] == "slug": r["pattern"] = "^[a-z0-9]+(?:-[a-z0-9]+)*$"
            if f["type"] == "date": r["format"] = "YYYY-MM-DD"
            if f["type"] == "images": r["multi_value_separator"] = "|"
            frules.append(r)
        rules["entities"][e["table"]] = {"id_prefix":e["id_prefix"],"fields":frules}
    rules["error_codes"] = {
        "REQUIRED_MISSING":"A required field is empty.",
        "FK_NOT_FOUND":"A linked ID does not exist in its parent sheet.",
        "NOT_IN_LIST":"Value is not one of the allowed drop-down options.",
        "DUPLICATE_ID":"This ID/slug already exists.",
        "SLUG_FORMAT":"Slug must be lowercase-with-dashes.",
        "BAD_NUMBER":"Field must be a number (no commas/text).",
        "BAD_DATE":"Date must be YYYY-MM-DD.",
        "OUT_OF_RANGE":"Number is below the allowed minimum.",
    }
    open(p("schema","validation_rules.json"),"w",encoding="utf-8").write(json.dumps(rules, ensure_ascii=False, indent=2))

# ---------------------------------------------------------------------------
if __name__ == "__main__":
    files = []
    files.append(build_master_workbook())
    for e in S.ENTITIES:
        files.append(build_entity_workbook(e))
    build_csv(); build_sql(); build_json_schema(); build_validation_rules()
    print("XLSX + CSV + schema written:")
    for f in files: print("  ", os.path.relpath(f, OUT))
    print("  schema/schema.sql, schema/schema.json, schema/validation_rules.json")
    print("  templates-csv/*.csv (+ __example.csv + _import_error_report_TEMPLATE.csv)")
